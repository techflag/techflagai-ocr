# -*- coding: utf-8 -*-
import sys 
import os  

project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

import json
import copy
from openpyxl import Workbook
# 移除旧的导入
# from table_structure_recognition.recover2excel import to_excel
# from common.box_util import stitch_boxes_into_lines_v2 as stitch_boxes_into_lines

# 新增：自定义的文本行拼接函数
def custom_stitch_boxes_into_lines(boxes, max_x_dist, min_y_overlap_ratio, symbol=""):
    """
    自定义的文本行拼接函数。
    将文本框按行拼接起来。
    返回拼接后的行列表，每行包含 'text', 'bbox', 'score', 'component_ids'。
    'component_ids' 是组成该行的原始 boxes 列表中的索引。
    """
    if not boxes:
        return []

    # 按 y 坐标（bbox的top），然后 x 坐标（bbox的left）排序
    # 每个 box 结构: {'bbox': [x1,y1,x2,y2], 'text': '...', 'original_idx': original_index_in_ocr_input, 'score': ...}
    sorted_boxes = sorted(boxes, key=lambda b: (b['bbox'][1], b['bbox'][0]))

    lines = []
    assigned_mask = [False] * len(sorted_boxes)

    for i in range(len(sorted_boxes)):
        if assigned_mask[i]:
            continue

        current_box = sorted_boxes[i]
        assigned_mask[i] = True

        current_line_text = current_box['text']
        current_line_bbox = list(current_box['bbox']) # 复制bbox
        current_line_components = [current_box['original_idx']]
        current_line_scores = [current_box.get('score', 1.0)] # 使用get获取分数，提供默认值

        # 尝试向右扩展当前行
        last_box_in_line_bbox = list(current_box['bbox'])

        for j in range(i + 1, len(sorted_boxes)):
            if assigned_mask[j]:
                continue

            candidate_box = sorted_boxes[j]

            # 1. 检查Y轴重叠率
            overlap_y = max(0, min(last_box_in_line_bbox[3], candidate_box['bbox'][3]) - max(last_box_in_line_bbox[1], candidate_box['bbox'][1]))
            min_height = min(last_box_in_line_bbox[3] - last_box_in_line_bbox[1], candidate_box['bbox'][3] - candidate_box['bbox'][1])
            y_overlap_ratio = overlap_y / min_height if min_height > 0 else 0
            
            if y_overlap_ratio >= min_y_overlap_ratio:
                # 2. 检查X轴距离 (候选框在当前行最后一个框的右边)
                # 确保 candidate_box['bbox'][0] (left) >= last_box_in_line_bbox[2] (right)
                # 并且距离小于 max_x_dist
                horizontal_distance = candidate_box['bbox'][0] - last_box_in_line_bbox[2]
                if 0 <= horizontal_distance < max_x_dist:
                    # 合并
                    current_line_text += symbol + candidate_box['text']
                    # 更新整行的bbox: x_min 不变, y_min 是所有组件的y_min的最小值, x_max 是新组件的x_max, y_max 是所有组件的y_max的最大值
                    current_line_bbox[1] = min(current_line_bbox[1], candidate_box['bbox'][1])
                    current_line_bbox[2] = candidate_box['bbox'][2] # x_max 更新为新组件的 x_max
                    current_line_bbox[3] = max(current_line_bbox[3], candidate_box['bbox'][3])
                    
                    current_line_components.append(candidate_box['original_idx'])
                    current_line_scores.append(candidate_box.get('score', 1.0))
                    assigned_mask[j] = True
                    last_box_in_line_bbox = list(candidate_box['bbox']) # 更新用于下次比较的最后一个框
                # else: # 如果X距离不满足，对于已排序的列表，后续的框X距离会更大，可以考虑break优化，但简单起见先不加
            # else: # 如果Y重叠不满足，对于已按Y排序的列表，如果当前行和候选框Y不重叠，后续的更不可能（除非行高变化很大）
                # 简单的贪心策略：只向右拼接在Y重叠范围内的最近的框

        avg_score = sum(current_line_scores) / len(current_line_scores) if current_line_scores else 1.0
        lines.append({
            'text': current_line_text,
            'bbox': current_line_bbox,
            'score': avg_score,
            'component_ids': current_line_components
        })
    return lines

# 新增：自定义的 Excel 生成函数
def custom_to_excel(data_list, workbook):
    """
    自定义的 Excel 生成函数。
    将处理后的数据写入 Workbook 对象。
    data_list 中的每个元素是一个字典，包含 'text', 'row' ([start_row, end_row]), 'col' ([start_col, end_col])。
    行和列索引是0-based。
    """
    if not data_list:
        if not workbook.sheetnames:
            workbook.create_sheet("Sheet1")
        return workbook

    sheet_name = "Sheet1"
    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(sheet_name)
    else:
        sheet = workbook[sheet_name]

    def ranges_overlap(r1_start, r1_end, r2_start, r2_end):
        # 检查一维范围是否重叠
        return not (r1_end < r2_start or r1_start > r2_end)

    def boxes_overlap(box1_sr, box1_sc, box1_er, box1_ec, box2_sr, box2_sc, box2_er, box2_ec):
        # 检查两个矩形框是否重叠 (坐标是1-based)
        overlap_row = ranges_overlap(box1_sr, box1_er, box2_sr, box2_er)
        overlap_col = ranges_overlap(box1_sc, box1_ec, box2_sc, box2_ec)
        return overlap_row and overlap_col

    for item in data_list:
        text = item.get('text', '')
        row_span = item.get('row')
        col_span = item.get('col')

        if not row_span or not col_span or len(row_span) != 2 or len(col_span) != 2:
            print(f"警告: 数据项 '{text[:20]}...' 的行或列范围无效，已跳过。行: {row_span}, 列: {col_span}")
            continue

        # openpyxl 使用 1-based 索引
        # item中的row/col是0-based
        item_start_r, item_end_r = row_span[0] + 1, row_span[1] + 1
        item_start_c, item_end_c = col_span[0] + 1, col_span[1] + 1
        
        # ADD THIS DEBUG BLOCK START
        if row_span == [0,0] and col_span == [0,0]: # Specifically for the A1 cell item
            print("-" * 40)
            print(f"DEBUG A1_ITEM: Encountered item for A1.")
            print(f"DEBUG A1_ITEM: Text to write = '{text}'")
            print(f"DEBUG A1_ITEM: Raw item row_span = {row_span}, col_span = {col_span}")
            print(f"DEBUG A1_ITEM: Target Excel cells: R={item_start_r}-{item_end_r}, C={item_start_c}-{item_end_c}")
            print(f"DEBUG A1_ITEM: Current merged cells before processing this item: {list(sheet.merged_cells.ranges)}")
        # ADD THIS DEBUG BLOCK END

        if item_start_r > item_end_r or item_start_c > item_end_c:
            print(f"警告: 数据项 '{text[:20]}...' 的行列范围无效 (start > end)，已跳过。行: {item_start_r}-{item_end_r}, 列: {item_start_c}-{item_end_c}")
            continue
        
        # 在写入和合并之前，解除与当前单元格区域重叠的任何现有合并
        merges_to_remove_coords = []
        # 迭代现有合并区域的副本，因为我们可能会修改它
        for mc in list(sheet.merged_cells.ranges):
            mc_s_r, mc_s_c, mc_e_r, mc_e_c = mc.min_row, mc.min_col, mc.max_row, mc.max_col
            if boxes_overlap(item_start_r, item_start_c, item_end_r, item_end_c, 
                             mc_s_r, mc_s_c, mc_e_r, mc_e_c):
                merges_to_remove_coords.append((mc_s_r, mc_s_c, mc_e_r, mc_e_c))
        
        for r_s, c_s, r_e, c_e in merges_to_remove_coords:
            try:
                sheet.unmerge_cells(start_row=r_s, start_column=c_s, end_row=r_e, end_column=c_e)
                # ADD THIS DEBUG PRINT
                if row_span == [0,0] and col_span == [0,0] and r_s == 1 and c_s == 1:
                     print(f"DEBUG A1_ITEM: Successfully unmerged cells involving A1: {r_s},{c_s}-{r_e},{c_e}")
            except KeyError:
                 print(f"信息: 尝试解除合并区域 {r_s},{c_s}-{r_e},{c_e} 时可能遇到问题，或该区域已被解除。")
            except Exception as unmerge_ex:
                print(f"警告: 解除合并区域 {r_s},{c_s}-{r_e},{c_e} 失败: {unmerge_ex}")


        # 写入文本到起始单元格
        # 此时 (item_start_r, item_start_c) 应该是一个常规单元格
        try:
            sheet.cell(row=item_start_r, column=item_start_c, value=text)
            # ADD THIS DEBUG PRINT
            if row_span == [0,0] and col_span == [0,0]:
                print(f"DEBUG A1_ITEM: Successfully wrote '{text}' to cell ({item_start_r},{item_start_c})")
                print("-" * 40)
        except Exception as cell_write_ex:
            print(f"错误: 写入单元格 ({item_start_r},{item_start_c}) 失败，内容 '{text[:20]}...'. 错误: {cell_write_ex}")
            continue # 跳过此项

        # 如果需要合并单元格 (即当前项定义的区域大于1x1)
        if item_start_r != item_end_r or item_start_c != item_end_c:
            try:
                sheet.merge_cells(start_row=item_start_r, start_column=item_start_c, end_row=item_end_r, end_column=item_end_c)
            except Exception as e:
                print(f"错误: 合并单元格失败 {item_start_r},{item_start_c} 到 {item_end_r},{item_end_c} 内容 '{text[:20]}...'. 错误: {e}")
    return workbook

def parse_ocr_data_to_excel(ocr_data_path, output_excel_path):
    try:
        with open(ocr_data_path, 'r', encoding='utf-8') as f:
            loaded_data = json.load(f)
    except FileNotFoundError:
        print(f"错误: 输入文件 {ocr_data_path} 未找到。")
        workbook = Workbook()
        if not workbook.sheetnames: workbook.create_sheet("Sheet1")
        workbook.save(output_excel_path)
        print(f"已保存一个空的 Excel 文件到: {output_excel_path} (由于输入文件未找到)")
        return
    except json.JSONDecodeError:
        print(f"错误: 解析JSON文件 {ocr_data_path} 失败。")
        workbook = Workbook()
        if not workbook.sheetnames: workbook.create_sheet("Sheet1")
        workbook.save(output_excel_path)
        print(f"已保存一个空的 Excel 文件到: {output_excel_path} (由于JSON解析错误)")
        return

    cor_like_output = []

    if isinstance(loaded_data, dict):
        print("信息: 输入的JSON是一个字典。将按 'stitched_lines' 和 'original_info' 结构处理。")
        stitched_lines = loaded_data.get('stitched_lines', [])
        original_components_info = loaded_data.get('original_info', [])

        if not stitched_lines:
            print("警告: 在字典格式的JSON中未找到 'stitched_lines' 或其为空。")
        if not original_components_info and stitched_lines: # Only warn if stitched_lines exist but original_info is missing
            print("警告: 在字典格式的JSON中未找到 'original_info' 或其为空，这可能导致无法处理 'stitched_lines'。")

        for i, stitched_line_info in enumerate(stitched_lines):
            if not isinstance(stitched_line_info, dict):
                print(f"警告: 拼接行 {i} 不是一个字典。已跳过此行。")
                continue

            if 'component_ids' not in stitched_line_info:
                print(f"警告: 拼接行 {i} (文本: '{stitched_line_info.get('text', '')[:30]}...') 缺少 'component_ids'。已跳过此行。")
                continue

            component_ids = stitched_line_info['component_ids']
            if not isinstance(component_ids, list):
                print(f"警告: 拼接行 {i} 的 'component_ids' 不是列表。已跳过此行。")
                continue
            
            min_r_start, max_r_end = float('inf'), float('-inf')
            min_c_start, max_c_end = float('inf'), float('-inf')
            found_valid_original_info = False
            
            processed_box = stitched_line_info.get('box')

            for component_idx_in_ocr_input in component_ids:
                if not (isinstance(component_idx_in_ocr_input, int) and 
                        0 <= component_idx_in_ocr_input < len(original_components_info)):
                    print(f"警告: 组件索引 {component_idx_in_ocr_input} 无效或超出范围 (原始组件数量: {len(original_components_info)})。已跳过此组件。")
                    continue
                
                original_item = original_components_info[component_idx_in_ocr_input]
                if not isinstance(original_item, dict):
                    print(f"警告: 原始组件 {component_idx_in_ocr_input} 不是一个字典。已跳过此组件。")
                    continue
                
                has_original_row = 'row' in original_item and isinstance(original_item['row'], list) and len(original_item['row']) == 2
                has_original_col = 'col' in original_item and isinstance(original_item['col'], list) and len(original_item['col']) == 2

                if has_original_row:
                    try:
                        min_r_start = min(min_r_start, int(original_item['row'][0]))
                        max_r_end = max(max_r_end, int(original_item['row'][1]))
                        found_valid_original_info = True
                    except (ValueError, TypeError):
                        print(f"警告: 组件 {component_idx_in_ocr_input} 的行信息 '{original_item['row']}' 包含无效值。")
                
                if has_original_col:
                    try:
                        min_c_start = min(min_c_start, int(original_item['col'][0]))
                        max_c_end = max(max_c_end, int(original_item['col'][1]))
                        found_valid_original_info = True
                    except (ValueError, TypeError):
                         print(f"警告: 组件 {component_idx_in_ocr_input} 的列信息 '{original_item['col']}' 包含无效值。")

            if not found_valid_original_info or min_r_start == float('inf') or min_c_start == float('inf'):
                print(f"警告: 无法为拼接行 {i} (文本: '{stitched_line_info.get('text', '')[:30]}...') 确定有效的原始行列信息。已跳过此行。")
                continue

            actual_row_span = [min_r_start, max_r_end]
            actual_col_span = [min_c_start, max_c_end]
            
            item_to_add = {
                'row': actual_row_span,
                'col': actual_col_span,
                'box': processed_box, 
                'text': stitched_line_info.get('text',''), # Ensure text exists
                'score': float(stitched_line_info.get('score', 1.0)), 
                'type': 'text_stitched' 
            }
            cor_like_output.append(item_to_add)

    elif isinstance(loaded_data, list):
        print("信息: 输入的JSON是一个列表。将尝试直接处理列表中的项目。")
        for i, item in enumerate(loaded_data):
            if not isinstance(item, dict):
                print(f"警告: 列表中的项目索引 {i} 不是一个字典。已跳过。")
                continue
            
            # 步骤 1: 强制检查 'row' 和 'col' 的存在性和基本列表格式
            if not ('row' in item and isinstance(item['row'], list) and len(item['row']) == 2 and
                    'col' in item and isinstance(item['col'], list) and len(item['col']) == 2):
                item_text_for_log = str(item.get('text', 'N/A'))[:20] # 用于日志，如果text存在
                print(f"警告: 列表中的项目索引 {i} (文本片段: '{item_text_for_log}...') 缺少 'row' 或 'col' 键，或其基本格式不正确 (非列表或长度不为2)。无法定位单元格。已跳过。")
                print(f"  - Row: {item.get('row')}, Col: {item.get('col')}")
                continue
            
            # 步骤 2: 验证 'row' 和 'col' 内部值的有效性 (可转换为整数，且 start <= end)
            try:
                current_row_span = [int(r) for r in item['row']]
                current_col_span = [int(c) for c in item['col']]
                if not (current_row_span[0] <= current_row_span[1] and current_col_span[0] <= current_col_span[1]):
                    item_text_for_log = str(item.get('text', 'N/A'))[:20]
                    print(f"警告: 列表中的项目索引 {i} (文本片段: '{item_text_for_log}...') 的行列范围无效 (start > end)。行: {current_row_span}, 列: {current_col_span}。已跳过。")
                    continue
            except (ValueError, TypeError):
                item_text_for_log = str(item.get('text', 'N/A'))[:20]
                print(f"警告: 列表中的项目索引 {i} (文本片段: '{item_text_for_log}...') 的 'row' 或 'col' 包含无法转换为整数的值。已跳过。")
                continue

            # 步骤 3: 获取文本内容，如果缺失或为null则默认为空字符串
            cell_text_value = item.get('text') # 获取text字段的值，可能为None
            if cell_text_value is None: # 处理 'text' 键不存在或值为 null 的情况
                cell_text = ''
            else:
                cell_text = str(cell_text_value) # 确保文本是字符串类型

            item_to_add = {
                'row': current_row_span,
                'col': current_col_span,
                'text': cell_text, # 使用处理过的 cell_text
                'box': item.get('box'), 
                'score': float(item.get('score', 1.0)), 
                'type': item.get('type', 'text') 
            }
            cor_like_output.append(item_to_add)
        
        if not cor_like_output and loaded_data: # 如果列表非空但处理后输出为空
             print("警告: 输入JSON列表已处理，但没有项目符合写入Excel的格式要求（有效的row/col）。")

    else:
        print(f"错误: OCR 数据文件 {ocr_data_path} 的内容既不是有效的JSON对象也不是JSON列表。")
        workbook = Workbook()
        if not workbook.sheetnames: workbook.create_sheet("Sheet1")
        workbook.save(output_excel_path)
        print(f"已保存一个空的 Excel 文件到: {output_excel_path} (由于输入数据格式不支持)")
        return

    if not cor_like_output:
        print("没有有效数据可写入 Excel。这可能是因为：\n"
              "1. 如果输入是字典格式：所有拼接后的文本行都缺少 'component_ids' 或其组件缺少行列信息。\n"
              "2. 如果输入是列表格式：列表中的项目不符合预期的格式（缺少 'row', 'col', 'text' 或格式错误）。\n"
              "3. 输入文件本身为空或格式完全错误。")
        workbook = Workbook() 
        if not workbook.sheetnames:
            workbook.create_sheet("Sheet1")
        workbook.save(output_excel_path)
        print(f"已保存一个包含默认空工作表的 Excel 文件到: {output_excel_path}")
        return

    workbook = Workbook()
    if workbook.active and len(workbook.sheetnames) == 1 and workbook.active.title == "Sheet":
         workbook.remove(workbook.active)

    workbook_processed = custom_to_excel(cor_like_output, workbook=workbook)

    if workbook_processed is not None:
        if not workbook_processed.sheetnames: 
            print(f"错误: 'custom_to_excel' 函数返回的 workbook 对象不包含任何工作表。将保存一个包含默认空工作表的 Excel 文件。")
            workbook_processed.create_sheet("Sheet1") 
        workbook_processed.save(output_excel_path)
        print(f"处理后的数据已保存到 Excel 文件: {output_excel_path}")
    else:
        print(f"错误: 'custom_to_excel' 函数未能返回有效的 workbook 对象。将保存一个包含默认空工作表的 Excel 文件。")
        fallback_workbook = Workbook() 
        if not fallback_workbook.sheetnames:
            fallback_workbook.create_sheet("Sheet1")
        fallback_workbook.save(output_excel_path)
        print(f"已保存一个包含默认空工作表的 Excel 文件到: {output_excel_path} (作为回退)")

if __name__ == '__main__':
    input_json_file = '/Volumes/Work/java_work/common_table/test/test_table.json'
    output_excel_file = '/Volumes/Work/java_work/common_table/test/parsed_test_table.xlsx'
    
    os.makedirs(os.path.dirname(output_excel_file), exist_ok=True)
    parse_ocr_data_to_excel(input_json_file, output_excel_file)
